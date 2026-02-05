from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import csv
import time

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://www.sharesansar.com/agm-list")

wait = WebDriverWait(driver, 10)
data = []

while True:
    # Wait for table to load
    wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "table.table-bordered.table-striped")))
    time.sleep(1)

    # Scrape rows
    rows = driver.find_elements(
        By.CSS_SELECTOR, "table.table-bordered.table-striped tr")
    for row in rows[1:]:  # skip header
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) >= 4:
            data.append({
                "SNO": cols[0].text.strip(),
                "SYMBOL": cols[1].text.strip(),
                "COMPANY": cols[2].text.strip(),
                "AGM": cols[3].text.strip()
            })

   # 3️⃣ Check Next button
    next_btn = wait.until(EC.element_to_be_clickable((By.ID, "myTableC_next")))
    if "disabled" in next_btn.get_attribute("class"):
        break  # last page reached
    else:
        next_btn.click()
        time.sleep(2)  # wait for table to reload

driver.quit()

# # Save all rows to CSV
# csv_file = "agm_list_all_pages.csv"
# with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
#     writer = csv.DictWriter(f, fieldnames=["SNO", "SYMBOL", "COMPANY", "AGM"])
#     writer.writeheader()
#     writer.writerows(data)

# print(f"Saved {len(data)} rows from all pages to {csv_file}")

#  --- 2️⃣ Create Excel file ---
wb = Workbook()
ws = wb.active
ws.title = "AGM List"

# Headers
headers = ["SNO", "SYMBOL", "COMPANY", "AGM"]
ws.append(headers)

# Add data rows
for row in data:
    ws.append(list(row.values()))

# --- 3️⃣ Apply borders ---
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# --- 4️⃣ Auto-fit columns ---
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # get column name
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = max_length + 2  # padding
    ws.column_dimensions[column].width = adjusted_width

# --- 5️⃣ Save Excel ---
excel_file = "agm_list.xlsx"
wb.save(excel_file)
print(f"Saved {len(data)} rows to {excel_file}")
